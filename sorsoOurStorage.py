import os
import math
from openpyxl import load_workbook
from datetime import datetime

def take_products_from_stock(order, stock_file: str) -> dict:
    
    # Открытие файлов склада и шаблона заказа
    stock_wb = load_workbook(stock_file)
    stock_ws = stock_wb.active

    order_wb = load_workbook(order_template_file)
    order_ws = order_wb.active

    # Столбцы
    stock_name_col = 'A'  # Артикул на складе
    stock_amount_col = 'B'  # Количество на складе

    # order_name_col = 'G'  # Артикул в заказе
    # order_bundle_col = 'I'  # Фасовка в заказе
    # order_amount_col = 'L'  # Количество штук в заказе (перезаписывается)

    # создаем dict для данных
    stock_data = {}

    for row in range(2, stock_ws.max_row + 1):
        article = str(stock_ws[f'{stock_name_col}{row}'].value) 
        amount = stock_ws[f'{stock_amount_col}{row}'].value or 0

        if article:
            stock_data[article] = {
                'amount': int(amount), 
                'row': row
            }
       
    # словарь для создания заказа
    updated_order = {}

   # Обработка заказа из словаря
    for article, (bundle, order_quantity) in order.items():
        if not article or not bundle or not order_quantity:
            continue

        bundle = int(bundle)
        order_quantity = int(order_quantity)

        if article in stock_data:
            
            # Данные о товаре на складе
            stock_amount = stock_data[article]['amount']
            stock_row = stock_data[article]['row']

            remaining_items = max(order_quantity - stock_amount, 0) # Считаем количество товара, который нужно дозаказать
            order_packs = math.ceil(remaining_items / bundle) # считаем кол-во пачек для заказа 
            
            # Если на складе хватает товара
            if stock_amount >= order_quantity: 
                stock_amount = stock_amount - order_quantity # Уменьшаем количество на складе
                stock_ws[f'{stock_amount_col}{stock_row}'].value = stock_amount # Обновляем склад
              
                # order_quantity = None  # Устанавливаем None (пустую ячейку) для количества в заказе
                updated_order[article] = None  # Обновляем словарь с заказом

            # Если на складе товара недостаточно
            else:
                stock_amount = order_packs * bundle + stock_amount - order_quantity  # Все, что было на складе, используется
                stock_ws[f'{stock_amount_col}{stock_row}'].value = stock_amount  # Обновляем склад
                
                # order_quantity = order_packs # Записываем количество упаковок в заказ
                updated_order[article] = order_packs  # Обновляем словарь с заказом
        else:
            # Если артикула нет на складе, считаем полное количество упаковок
            order_packs = math.ceil(order_quantity / bundle)
            # order_quantity = order_packs  # Записываем количество упаковок
            updated_order[article] = order_packs  # Обновляем словарь с заказом
            
    # Сохранить изменения
    stock_wb.save(stock_file)
    stock_wb.close()

    return updated_order

def write_order_to_template(order: dict, order_template_file: str, result_dir: str):
    order_wb = load_workbook(order_template_file)
    order_ws = order_wb.active

    order_name_col = 'G'
    order_packs_col = 'L'

    for row in range(10, order_ws.max_row + 1):
        article = str(order_ws[f"{order_name_col}{row}"].value)
        if article in order:
            order_ws[f"{order_packs_col}{row}"] = order[article]

    now = datetime.now().strftime("%d.%m.%Y %H-%M")
    result_file = os.path.join(result_dir, f"Заказ Sorso {now}.xlsx")
    order_wb.save(result_file)
    order_wb.close()

    return result_file

if __name__ == "__main__":
    stock_folder = 'Stock'
    order_folder = 'Order template'
    result_dir = 'Results'

    def find_xlsx_file(folder_path):
        for file in os.listdir(folder_path):
            if file.endswith(".xlsx"):
                return os.path.join(folder_path, file)
        return None

    stock_file = find_xlsx_file(stock_folder)
    order_template_file = find_xlsx_file(order_folder)    
    
    if stock_file and order_template_file:
        # Заполнение словаря order из таблицы order_template_file
        order_wb = load_workbook(order_template_file)
        order_ws = order_wb.active

        order_name_col = 'G'
        order_bundle_col = 'I'
        order_amount_col = 'L'

        order = {}
        for row in range(10, order_ws.max_row + 1):
            article = str(order_ws[f"{order_name_col}{row}"].value)
            bundle = order_ws[f"{order_bundle_col}{row}"].value
            order_quantity = order_ws[f"{order_amount_col}{row}"].value

            if article and bundle and order_quantity:
                order[article] = (int(bundle), int(order_quantity))

        order_wb.close()
    
        updated_order = take_products_from_stock(order, stock_file)
        result_path = write_order_to_template(updated_order, order_template_file, result_dir)
        print(f"Обновленный заказ сохранен в: {result_path}")
    else:
        print("Файлы не найдены.")
